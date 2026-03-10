
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from datetime import datetime , date
import matplotlib.pyplot as plt
import os
import re

# -----------------------------
# FILE PATHS
# -----------------------------
CUSTOMERS_FILE = "customers.xlsx"
PRODUCTS_FILE = "products.xlsx"
ORDERS_FILE = "orders.xlsx"

# -----------------------------
# DATABASE CONNECTION
# -----------------------------
DB_USER = input("Enter Oracle DB username: ")
DB_PASSWORD = input("Enter Oracle DB password: ")
DB_HOST = input("Enter DB host (default localhost): ") or "localhost"
DB_PORT = input("Enter DB port (default 1522): ") or "1522"
DB_SERVICE = input("Enter DB service name (default XEXDB): ") or "XEXDB"


engine = create_engine(
    f"oracle+oracledb://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/?service_name={DB_SERVICE}"
)
print("Database connected successfully")

# -----------------------------
# HELPER FUNCTIONS
# -----------------------------

def is_row_empty(row, cols):
    """
    Check if all specified columns in the row are empty, None, or 'Unknown'.
    Returns True if all are empty.
    """
    for col in cols:
        val = row.get(col)
        if val is not None and str(val).strip() not in ["", "Unknown"]:
            return False
    return True

from openpyxl.utils import get_column_letter

def auto_adjust_excel_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2
    wb.save(file_path)

def advanced_clean_row(row, numeric_cols=[], string_cols=[], key_cols=[], text_case=None):
    # Numeric columns
    for col in numeric_cols:
        if col in row:

            try:
                row[col] = float(row[col])

                if row[col] < 0:
                   row[col] = 0
            except:
                row[col] = 0
  

    # String columns
    for col in string_cols:
        if col in row:
            if row[col] is None or str(row[col]).strip() == "":
                row[col] = "Unknown"
            else:
                row[col] = str(row[col]).strip()
                if text_case == "title":
                    row[col] = row[col].title()
                elif text_case == "upper":
                    row[col] = row[col].upper()
                elif text_case == "lower":
                    row[col] = row[col].lower()

    # Key columns 
    for col in key_cols:
        if col in row and (row[col] is None or pd.isna(row[col])):
            row[col] = 0

    return row

def save_to_db_append_or_update(df, table_name, key_column):
    df = df.copy()
    # Convert numeric types to int for Oracle
    for col in df.select_dtypes(include=["int64", "float64"]).columns:
        df[col] = df[col].astype(int)
    # Convert datetime properly
    for col in df.select_dtypes(include=["datetime64", "datetime"]).columns:
        df[col] = pd.to_datetime(df[col])

    try:
        existing_df = pd.read_sql(f"SELECT * FROM {table_name}", engine)
        new_rows = df[~df[key_column].isin(existing_df[key_column])]
        if not new_rows.empty:
            new_rows.to_sql(table_name, engine, if_exists="append", index=False)
            print(f"{len(new_rows)} new rows inserted into {table_name}")
        else:
            print(f"No new rows to insert into {table_name}")
    except Exception as e:
    # Log error to file
        with open("db_error.log", "a") as log_file:
           log_file.write(f"{datetime.now()}: Error updating {table_name} - {e}\n")
    # Fallback to Excel/DB
        df.to_sql(table_name, engine, if_exists="replace", index=False)
        print(f"{table_name} table created. (Original error: {e})")

# -----------------------------
# CUSTOMER SIGNUP / LOGIN (GET ACTUAL DATA FROM SIGNUP PAGE)
# -----------------------------
def register_customer(name, email, address, phone):
    expected_columns = ["customer_id", "name", "email", "address", "phone", "signup_date", "last_login"]

    if os.path.exists(CUSTOMERS_FILE):
        df = pd.read_excel(CUSTOMERS_FILE)
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    else:
        df = pd.DataFrame(columns=expected_columns)
    
    # ---------------------- Check if customer already exists ----------------------
    existing_customer = df[df['email'].str.lower() == email.lower()]

    if not existing_customer.empty:
    # Update all relevant fields for returning customer
        df.loc[df['email'].str.lower() == email.lower(), ['name','address','phone','last_login']] = [
        name.title(), address.title(), int(''.join(filter(str.isdigit, phone))), pd.to_datetime(datetime.now().date())
    ]
    
    # Ensure signup_date remains as datetime
        df['signup_date'] = pd.to_datetime(df['signup_date'])

    # Save to Excel and DB
        df.to_excel(CUSTOMERS_FILE, index=False, engine="openpyxl")
        auto_adjust_excel_columns(CUSTOMERS_FILE)
        save_to_db_append_or_update(df, "customers", "customer_id")

        print(f"Returning customer updated: {existing_customer['customer_id'].values[0]}")
        return existing_customer['customer_id'].values[0]



    # Generate unique customer_id using Excel + DB
    try:
        db_customers = pd.read_sql("SELECT MAX(customer_id) AS max_id FROM customers", engine)
        max_db_id = db_customers["max_id"][0] if db_customers["max_id"][0] is not None else 1000
    except:
        max_db_id = 1000

    max_excel_id = df["customer_id"].max() if not df.empty else 1000
    new_id = max(max_db_id, max_excel_id) + 1

    today = date.today()
    new_row = {
        "customer_id": new_id,
        "name": name,
        "email": email,
        "address": address,
        "phone": phone,
        "signup_date": today,
        "last_login": today
    }

    new_row = advanced_clean_row(
        new_row,
        numeric_cols=["customer_id"],
        string_cols=["name","email","address","phone"],
        key_cols=["customer_id"],
        text_case="title"
    )
    # ---------------------- Email Validation ----------------------
    email_pattern = r"^[\w\.-]+@[\w\.-]+\.\w+$"
    if not re.match(email_pattern, new_row["email"]):
        print("Error: Invalid email format.")
        return None

    # ---------------------- Phone Validation ----------------------
    phone_digits = ''.join(filter(str.isdigit, new_row["phone"]))
    if len(phone_digits) < 10 or len(phone_digits) > 12:
        print("Error: Invalid phone number.")
        return None
    else:
        new_row["phone"] = int(phone_digits)   

    # --- Skip completely empty rows ---
    if is_row_empty(new_row, ["name", "email", "address", "phone"]):
       print("Error: Cannot register customer with all empty fields.")
       return None

    df.loc[len(df)] = new_row
    df.to_excel(CUSTOMERS_FILE, index=False, engine="openpyxl")
    auto_adjust_excel_columns(CUSTOMERS_FILE)
    save_to_db_append_or_update(df, "customers", "customer_id")
    print(f"Customer registered: {new_id}")
    return new_id

# -----------------------------
# ADD / UPDATE PRODUCT (GET ACTUAL DATA FROMA DMIN PANEL )
# -----------------------------
def add_or_update_product(product_id, product_name, price, category, supplier, stock=None):
    expected_columns = ["product_id","product_name","price","stock","category","supplier"]

    if os.path.exists(PRODUCTS_FILE):
        df = pd.read_excel(PRODUCTS_FILE)
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    else:
        df = pd.DataFrame(columns=expected_columns)

    new_row = {
        "product_id": product_id,
        "product_name": product_name,
        "price": price,
        "stock": stock if stock is not None else 0,
        "category": category,
        "supplier": supplier
    }

    new_row = advanced_clean_row(
        new_row,
        numeric_cols=["product_id","price","stock"],
        string_cols=["product_name","category","supplier"],
        key_cols=["product_id"],
        text_case="title"
    )

    if is_row_empty(new_row, ["product_name","price","category","supplier"]):
        print("Error: Cannot add product with all empty fields.")
        return

    # Check if product exists
    existing = df[df["product_id"] == new_row["product_id"]]
    if not existing.empty:
        # Update all fields in one step
        df.loc[df["product_id"] == new_row["product_id"], ["product_name","price","category","supplier","stock"]] = [
            new_row["product_name"], new_row["price"], new_row["category"], new_row["supplier"], new_row["stock"]
        ]
        print(f"Product updated: {product_id}")

        # --- Cascade price change to orders ---
        if os.path.exists(ORDERS_FILE):
            orders_df = pd.read_excel(ORDERS_FILE)
            orders_df.columns = orders_df.columns.str.strip().str.lower().str.replace(" ", "_")
            mask = orders_df["product_id"] == product_id
            if mask.any():
                orders_df.loc[mask, "price"] = orders_df.loc[mask, "quantity"] * new_row["price"]
                orders_df.to_excel(ORDERS_FILE, index=False, engine="openpyxl")
                auto_adjust_excel_columns(ORDERS_FILE)
                save_to_db_append_or_update(orders_df, "orders", "order_id")
                print(f"Updated prices in orders for product {product_id}")
    else:
        df.loc[len(df)] = new_row
        print(f"New product added: {product_id}")

    # Save product changes
    df.to_excel(PRODUCTS_FILE, index=False, engine="openpyxl")
    auto_adjust_excel_columns(PRODUCTS_FILE)
    save_to_db_append_or_update(df, "products", "product_id")


# -----------------------------
# REGISTER ORDER ( GET ACTUAL DATA AT CHECKOUT PAGE )
# -----------------------------
def register_order(customer_id, product_id, quantity, payment_method, price):
    expected_columns = ["order_id","customer_id","product_id","quantity","order_date","payment_method","price"]

    if os.path.exists(ORDERS_FILE):
        df = pd.read_excel(ORDERS_FILE)
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    else:
        df = pd.DataFrame(columns=expected_columns)

    now = datetime.now()

    # --- Get product price from PRODUCTS_FILE ---
    products_df_temp = pd.read_excel(PRODUCTS_FILE)
    products_df_temp.columns = products_df_temp.columns.str.strip().str.lower().str.replace(" ", "_")
    product_row = products_df_temp[products_df_temp["product_id"] == product_id]
    if product_row.empty:
        print(f"Error: Product {product_id} not found")
        return
    product_price = product_row["price"].values[0]

    # Check if order already exists (by order_id or by customer+product combo)
    existing_order = df[(df["customer_id"] == customer_id) & (df["product_id"] == product_id)]

    if not existing_order.empty:
        # Update existing order with automatic price calculation
        order_id = existing_order["order_id"].values[0]
        df.loc[df["order_id"] == order_id, ["quantity","payment_method","price","order_date"]] = [
            quantity, payment_method, product_price * quantity, now
        ]
        print(f"Order updated: {order_id}")
    else:
        # Generate new order_id from DB + Excel
        try:
            db_orders = pd.read_sql("SELECT MAX(order_id) AS max_id FROM orders", engine)
            max_db_id = db_orders["max_id"][0] if db_orders["max_id"][0] is not None else 5000
        except:
            max_db_id = 5000

        max_excel_id = df["order_id"].max() if not df.empty else 5000
        order_id = max(max_db_id, max_excel_id) + 1

        # Calculate total price automatically
        total_price = product_price * quantity

        new_row = {
            "order_id": order_id,
            "customer_id": customer_id,
            "product_id": product_id,
            "quantity": quantity,
            "order_date": now,
            "payment_method": payment_method,
            "price": total_price
        }

        new_row = advanced_clean_row(
            new_row,
            numeric_cols=["order_id","customer_id","product_id","quantity","price"],
            string_cols=["payment_method"],
            key_cols=["order_id"]
        )

        if is_row_empty(new_row, ["customer_id","product_id","quantity","price"]):
            print("Error: Cannot register order with missing data.")
            return

        df.loc[len(df)] = new_row
        print(f"Order registered: {order_id}")

    # Save orders to Excel and DB
    df.to_excel(ORDERS_FILE, index=False, engine="openpyxl")
    auto_adjust_excel_columns(ORDERS_FILE)
    save_to_db_append_or_update(df, "orders", "order_id")

    # Update stock
    products_df = pd.read_excel(PRODUCTS_FILE)
    products_df.columns = products_df.columns.str.strip().str.lower().str.replace(" ", "_")

    product_stock = products_df.loc[products_df["product_id"] == product_id, "stock"].values[0]
    if quantity > product_stock:
        print(f"Error: Not enough stock for product {product_id}. Available: {product_stock}")
        return

    products_df.loc[products_df["product_id"] == product_id, "stock"] -= quantity

    products_df.to_excel(PRODUCTS_FILE, index=False, engine="openpyxl")
    auto_adjust_excel_columns(PRODUCTS_FILE)
    save_to_db_append_or_update(products_df, "products", "product_id")
    print(f"Stock updated for product {product_id}")
# -----------------------------
# DEMO DATA FOR ANALYSIS 
# -----------------------------
def generate_demo_data_grocery():
    customers = [
        ("Ali Khan","ali@gmail.com","Karachi","03001234567"),
        ("Sara Ahmed","sara@gmail.com","Lahore","03111234567"),
        ("Usman Tariq","usman@gmail.com","Islamabad","03211234567"),
        ("Hina Malik","hina@gmail.com","Karachi","03311234567"),
        ("Bilal Sheikh","bilal@gmail.com","Hyderabad","03411234567"),
        ("Fatima Noor","fatima@gmail.com","Lahore","03511234567"),
        ("Ahmed Raza","ahmed@gmail.com","Karachi","03611234567"),
        ("Zara Khan","zara@gmail.com","Multan","03711234567"),
        ("Hamza Ali","hamza@gmail.com","Faisalabad","03811234567"),
        ("Ayesha Malik","ayesha@gmail.com","Karachi","03911234567"),
        ("Imran Shah","imran@gmail.com","Peshawar","03021234567"),
        ("Noor Fatima","noor@gmail.com","Lahore","03121234567"),
        ("Saad Ahmed","saad@gmail.com","Karachi","03221234567"),
        ("Sana Khan","sana@gmail.com","Quetta","03321234567"),
        ("Danish Ali","danish@gmail.com","Multan","03421234567")
    ]

    customer_ids = []
    for name,email,address,phone in customers:
        cid = register_customer(name,email,address,phone)
        customer_ids.append(cid)
    print("15 Customers inserted")

    products = [
        (101,"Milk 1L",260,"Dairy","DairyFarm",50),
        (102,"Eggs Dozen",300,"Dairy","FarmFresh",30),
        (103,"Bread Loaf",150,"Bakery","BakerHouse",40),
        (104,"Cheddar Cheese 200g",400,"Dairy","CheeseCo",25),
        (105,"Butter 250g",350,"Dairy","DairyFarm",20),
        (106,"Apple 1kg",500,"Fruits","FreshFarms",35),
        (107,"Banana 1kg",200,"Fruits","Tropico",50),
        (108,"Tomato 1kg",250,"Vegetables","VeggieLand",30),
        (109,"Potato 1kg",120,"Vegetables","VeggieLand",40),
        (110,"Onion 1kg",180,"Vegetables","VeggieLand",45),
        (111,"Rice 5kg",1200,"Grains","GrainWorld",15),
        (112,"Wheat Flour 5kg",1000,"Grains","GrainWorld",20),
        (113,"Sugar 1kg",150,"Pantry","SweetCo",25),
        (114,"Olive Oil 500ml",600,"Pantry","OliveCo",10),
        (115,"Tomato Sauce 500ml",300,"Pantry","Saucy",30)
    ]

    for p in products:
        add_or_update_product(*p)
    print("15 Grocery Products inserted")


    orders = [
        (customer_ids[0],101,2,"Cash"),        # Milk 1L × 2
        (customer_ids[1],102,1,"Credit Card"), # Eggs Dozen × 1
        (customer_ids[2],103,3,"Cash"),        # Bread Loaf × 3
        (customer_ids[3],104,2,"Debit Card"),  # Cheddar Cheese 200g × 2
        (customer_ids[4],105,1,"Cash"),        # Butter 250g × 1
        (customer_ids[5],106,4,"Credit Card"), # Apple 1kg × 4
        (customer_ids[6],107,3,"Cash"),        # Banana 1kg × 3
        (customer_ids[7],108,5,"Debit Card"),  # Tomato 1kg × 5
        (customer_ids[8],109,2,"Credit Card"), # Potato 1kg × 2
        (customer_ids[9],110,3,"Cash"),        # Onion 1kg × 3
        (customer_ids[10],111,1,"Credit Card"),# Rice 5kg × 1
        (customer_ids[11],112,2,"Cash"),       # Wheat Flour 5kg × 2
        (customer_ids[12],113,5,"Debit Card"), # Sugar 1kg × 5
        (customer_ids[13],114,1,"Credit Card"),# Olive Oil 500ml × 1
        (customer_ids[14],115,3,"Credit Card")        # Tomato Sauce 500ml × 3
    ]

    for order in orders:
        cid, pid, qty, payment = order
        register_order(cid, pid, qty, payment, price=None)

    print("15 Grocery Orders inserted")
    print("Grocery demo dataset generated successfully")



# -----------------------------
# 1 TOTAL SALES REVENUE
# -----------------------------
def total_sales():
    total = orders["price"].sum()
    print("Total Sales Revenue:", total)


# -----------------------------
# 2 MOST SOLD PRODUCT
# -----------------------------
def most_sold_product():
    product_sales = orders.groupby("product_id")["quantity"].sum().reset_index()
    merged = pd.merge(product_sales, products, on="product_id")

    top = merged.sort_values(by="quantity", ascending=False).head(5)

    print("\nTop Selling Products")
    print(top[["product_name","quantity"]])


# -----------------------------
# 3 REVENUE BY PRODUCT
# -----------------------------
def revenue_by_product():
    # 1️⃣ Sum total revenue per product
    revenue = orders.groupby("product_id")["price"].sum().reset_index()

    # 2️⃣ Merge with products to get product names
    merged = pd.merge(revenue, products, on="product_id", how="left")

    # 3️⃣ Rename the summed price column to avoid conflicts
    if "price_x" in merged.columns:
        merged.rename(columns={"price_x": "total_revenue"}, inplace=True)
    elif "price" in merged.columns:
        merged.rename(columns={"price": "total_revenue"}, inplace=True)
    else:
        # fallback: take first numeric column (just in case)
        for col in merged.columns:
            if pd.api.types.is_numeric_dtype(merged[col]):
                merged.rename(columns={col: "total_revenue"}, inplace=True)
                break

    # 4️⃣ Sort by total revenue
    merged = merged.sort_values(by="total_revenue", ascending=False)

    # 5️⃣ Print results
    print("\nRevenue by Product")
    print(merged[["product_name", "total_revenue"]])



    


# -----------------------------
# 4 TOP CUSTOMERS
# -----------------------------
def top_customers():
    spending = orders.groupby("customer_id")["price"].sum().reset_index()
    merged = pd.merge(spending, customers, on="customer_id")

    top = merged.sort_values(by="price", ascending=False).head(5)

    print("\nTop Customers")
    print(top[["name","price"]])


# -----------------------------
# 5 PAYMENT METHOD ANALYSIS
# -----------------------------
def payment_analysis():
    payment = orders["payment_method"].value_counts()

    print("\nPayment Method Distribution")
    print(payment)


# -----------------------------
# 6 SALES BY CATEGORY
# -----------------------------

def sales_by_category():
    # Merge orders with products to get category info
    merged = pd.merge(orders, products, on="product_id", how="left")

    # Determine which column has order price (usually price_x)
    if "price_x" in merged.columns:
        revenue_col = "price_x"
    elif "price" in merged.columns:
        revenue_col = "price"
    else:
        # fallback: pick the first numeric column
        numeric_cols = merged.select_dtypes(include="number").columns
        revenue_col = numeric_cols[0]

    # Group by category and sum order revenue
    category_sales = merged.groupby("category")[revenue_col].sum()

    print("\nSales by Category")
    print(category_sales)


# -----------------------------
# 7 LOW STOCK PRODUCTS
# -----------------------------
def low_stock_products():
    low_stock = products[products["stock"] < 5]

    print("\nLow Stock Products")
    print(low_stock[["product_name","stock"]])


# -----------------------------
# 8 CUSTOMER LOCATION ANALYSIS
# -----------------------------
def city_analysis():
    merged = pd.merge(orders, customers, on="customer_id")

    city_orders = merged["address"].value_counts()

    print("\nOrders by City")
    print(city_orders)
def sales_by_product_chart():
    # 1️⃣ Sum revenue per product
    revenue = orders.groupby("product_id")["price"].sum().reset_index()

    # 2️⃣ Merge with products to get names
    merged = pd.merge(revenue, products, on="product_id", how="left")

    # 3️⃣ Rename summed revenue column
    if "price_x" in merged.columns:
        merged.rename(columns={"price_x":"total_revenue"}, inplace=True)
    else:
        merged.rename(columns={"price":"total_revenue"}, inplace=True)

    # 4️⃣ Plot
    plt.figure(figsize=(10,6))
    plt.bar(merged["product_name"], merged["total_revenue"], color='skyblue')
    plt.title("Revenue by Product")
    plt.xlabel("Product")
    plt.ylabel("Revenue")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

def payment_method_chart():
    payment_counts = orders["payment_method"].value_counts()

    plt.figure(figsize=(6,6))
    plt.pie(payment_counts, labels=payment_counts.index, autopct="%1.1f%%", startangle=90, colors=["#66b3ff","#ff9999","#99ff99","#ffcc99"])
    plt.title("Payment Method Distribution")
    plt.tight_layout()
    plt.show()
def orders_by_city_chart():
    # Merge orders with customers safely
    merged = pd.merge(orders, customers[['customer_id', 'address']], on="customer_id", how="left")

    # Count orders per city
    city_orders = merged["address"].value_counts()

    # Plot bar chart
    plt.figure(figsize=(10,6))
    plt.bar(city_orders.index, city_orders.values, color='lightgreen')
    plt.title("Orders by City")
    plt.xlabel("City")
    plt.ylabel("Number of Orders")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

def category_sales_chart():
    # Merge orders with products safely
    merged = pd.merge(orders, products[['product_id', 'category']], on="product_id", how="left")

    # Group by category using orders' price
    category_sales = merged.groupby("category")["price"].sum()  # 'price' still comes from orders

    # Plot
    plt.figure(figsize=(8,6))
    plt.bar(category_sales.index, category_sales.values, color='orange')
    plt.title("Sales by Category")
    plt.xlabel("Category")
    plt.ylabel("Revenue")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()


if __name__ == "__main__":

    # 1️⃣ Generate demo data (even if Excel is empty)
    generate_demo_data_grocery()
    print("\nDemo data created successfully\n")

    # 2️⃣ Reload Excel files after demo insertion
    customers = pd.read_excel(CUSTOMERS_FILE)
    products = pd.read_excel(PRODUCTS_FILE)
    orders = pd.read_excel(ORDERS_FILE)
    print("Data loaded for analysis\n")

    # 3️⃣ Run all analysis
    total_sales()
    most_sold_product()
    revenue_by_product()
    top_customers()
    payment_analysis()
    sales_by_category()
    low_stock_products()
    city_analysis()

    # 4️⃣ Generate all charts
    sales_by_product_chart()
    payment_method_chart()
    orders_by_city_chart()
    category_sales_chart()
